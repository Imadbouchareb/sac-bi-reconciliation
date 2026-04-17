import streamlit as st
import pandas as pd
import re
import unicodedata
import io
from collections import defaultdict
from datetime import datetime

# ============================================================
#                    CONFIGURATION DE PAGE
# ============================================================
st.set_page_config(
    page_title="Réconciliation Objectifs PREV vs Power BI",
    layout="wide",
    page_icon="🎯",
    initial_sidebar_state="expanded"
)

pd.set_option('future.no_silent_downcasting', True)

# ============================================================
#                       STYLE CSS CUSTOM
# ============================================================
st.markdown("""
<style>
    .main-title {
        font-size: 2.1rem;
        font-weight: 700;
        color: #0f172a;
        margin-bottom: 0.25rem;
        letter-spacing: -0.02em;
    }
    .main-subtitle {
        color: #64748b;
        font-size: 0.95rem;
        margin-bottom: 1.5rem;
    }
    .main-subtitle b { color: #0f172a; }

    .kpi-card {
        background: #ffffff;
        padding: 1.25rem 1.5rem;
        border-radius: 12px;
        border: 1px solid #e2e8f0;
        border-left: 4px solid #8b5cf6;
        box-shadow: 0 1px 3px rgba(0,0,0,0.04);
        height: 100%;
    }
    .kpi-card-ok    { border-left-color: #16a34a; }
    .kpi-card-alert { border-left-color: #dc2626; }
    .kpi-label {
        color: #64748b;
        font-size: 0.78rem;
        text-transform: uppercase;
        letter-spacing: 0.06em;
        font-weight: 600;
        margin-bottom: 0.4rem;
    }
    .kpi-value {
        font-size: 2rem;
        font-weight: 700;
        color: #0f172a;
        line-height: 1;
    }
    .kpi-value-ok    { color: #15803d; }
    .kpi-value-alert { color: #b91c1c; }

    .upload-ok {
        padding: 0.4rem 0.7rem;
        background: #ede9fe;
        border-radius: 6px;
        color: #5b21b6;
        font-size: 0.8rem;
        margin-top: -0.5rem;
        margin-bottom: 0.8rem;
        font-weight: 500;
    }

    .diag-box {
        background: #f5f3ff;
        border-left: 3px solid #8b5cf6;
        padding: 0.7rem 1rem;
        border-radius: 6px;
        margin-bottom: 1rem;
        font-size: 0.88rem;
        color: #5b21b6;
    }

    .stTabs [data-baseweb="tab-list"] { gap: 6px; }
    .stTabs [data-baseweb="tab"] {
        padding: 8px 16px;
        border-radius: 8px 8px 0 0;
        font-weight: 500;
        font-size: 0.88rem;
    }

    .section-header {
        font-size: 1.1rem;
        font-weight: 600;
        color: #0f172a;
        margin: 1.5rem 0 0.75rem 0;
    }

    .sidebar-title {
        font-weight: 700;
        font-size: 1rem;
        color: #0f172a;
        margin-bottom: 0.5rem;
    }
</style>
""", unsafe_allow_html=True)


# ============================================================
#                    FONCTIONS UTILITAIRES
#              (identiques à l'application SAC)
# ============================================================
def safe_float_conversion(val):
    if isinstance(val, pd.Series): val = val.iloc[0]
    if pd.isna(val): return 0.0
    if isinstance(val, (int, float)): return float(val)
    s = str(val).strip()
    if s == '' or s.lower() in ['nan', 'nat', 'none', 'nc', '<na>']: return 0.0
    s = re.sub(r'[^\d,.-]', '', s)
    if '.' in s and ',' in s:
        if s.rfind(',') > s.rfind('.'): s = s.replace('.', '').replace(',', '.')
        else: s = s.replace(',', '')
    else:
        if s.count(',') > 1: s = s.replace(',', '')
        elif s.count('.') > 1: s = s.replace('.', '')
        else: s = s.replace(',', '.')
    if s in ['', '.', '-']: return 0.0
    try: return float(s)
    except ValueError: return 0.0


def clean_excel_text(text):
    if not isinstance(text, str): return text
    return re.sub(r'[-•=●◦○▪\x00-\x1F\x7F]', '', text).strip()


def remove_accents(input_str):
    if not isinstance(input_str, str): return ""
    nfkd_form = unicodedata.normalize('NFKD', input_str)
    return u"".join([c for c in nfkd_form if not unicodedata.combining(c)])


def generate_join_key(enseigne, rayon):
    enseigne_clean = remove_accents(str(enseigne)).upper()
    enseigne_compact = re.sub(r'[^A-Z0-9]', '', enseigne_clean)

    if "JEANCOUTU" in enseigne_compact: e = "JEANCO"
    elif "REDAPPLE" in enseigne_compact: e = "REDAPP"
    elif "BRUNET" in enseigne_compact: e = "BRUNET"
    elif "GAPOUTLET" in enseigne_compact: e = "GAPOUT"
    elif "GAP" in enseigne_compact: e = "GAP"
    elif "CORNERADF" in enseigne_compact: e = "CORNADF"
    elif "CORNERLOLLIPOPS" in enseigne_compact: e = "CORNLOL"
    elif "INTERMARCHEMOA" in enseigne_compact: e = "INTMOA"
    elif "INTERMARCHE" in enseigne_compact: e = "INTERM"
    elif "GHANTYMOA" in enseigne_compact: e = "GHAMOA"
    elif "GHANTY" in enseigne_compact: e = "GHANTY"
    else:
        words = re.findall(r'[A-Z0-9]+', enseigne_clean)
        e = "UNKNOWN"
        if words:
            e = words[0]
            if len(e) <= 3 and len(words) > 1: e += words[1]
        e = e[:6]

    r = re.sub(r'[^A-Z0-9]', '', remove_accents(str(rayon)).upper())
    return f"{e}_{r}"


# ============================================================
#                 MOTEUR DE RÈGLES MÉTIER
#              (identique à l'application SAC)
# ============================================================
def apply_business_rules(df, col_enseigne, report_type):
    df['Rayon'] = df['Rayon'].astype(str)

    def apply_rule(enseignes, rayons, new_rayon):
        ens_regex = '(' + '|'.join([re.escape(e.lower()) for e in enseignes]) + ')'
        ray_regex = '^(' + '|'.join([re.escape(r.lower().strip()) for r in rayons]) + ')$'
        mask = df[col_enseigne].astype(str).str.lower().str.contains(ens_regex, regex=True, na=False) & \
               df['Rayon'].str.lower().str.contains(ray_regex, regex=True, na=False)
        df.loc[mask, 'Rayon'] = new_rayon

    if report_type == "Diffusion":
        apply_rule(['El Corte'], ['ParCaiss', 'Cheveux'], 'Bijoux')
        apply_rule(['Conbipel'], ['Capsules'], 'Bijoux')
        apply_rule(['Upim'], ['Access', 'Cheveux', 'Lunettes'], 'Bijoux')
        apply_rule(['Grain de Malice'], ['Access'], 'Bijoux')
        apply_rule(['Hema'], ['Access', 'Actua', 'Mariage'], 'Lunettes')
        apply_rule(['GAP', 'GAP Outlet'], ['Access'], 'Bijoux')
        apply_rule(['Auchan'], ['Enfant'], 'Bijoux')
        apply_rule(['BZB'], ['Cheveux'], 'Bijoux')
        apply_rule(['Cache Cache'], ['Cheveux'], 'Bijoux')
        apply_rule(['Zebra'], ['Cheveux'], 'Bijoux')
        apply_rule(['Retail'], ['Beaute', 'Montres'], 'Access')
        apply_rule(['Lollipops Retail'], ['Montres'], 'Access')
        apply_rule(['El Corte'], ['Montres', 'Beaute'], 'Access')
        apply_rule(['GL', 'Galeries Lafayette'], ['Papeteri', 'Parfums'], 'Access')
        apply_rule(['Intermarché MOA', 'MOA'], ['ParCaiss'], 'Access')
        apply_rule(['GAP', 'GAP Outlet'], ['Acier', 'Deco'], 'Capsules')
        apply_rule(['GAP', 'GAP Outlet'], ['Catalog'], 'Enfant')
        apply_rule(['Corner ADF'], ['Access'], 'Deco')
        communes_map = {
            'ArGarant': 'Bijoux', 'Sacherie': 'Bijoux', 'Catalog': 'Lunettes de vue',
            'Papeteri': 'Cartes de voeux', 'Chaussur': 'Access', 'Mariage': 'Bijoux',
            'Actua': 'Bijoux', 'Montres': 'Bijoux', 'ParCaiss': 'Parcours Caisse',
            'PARCAISS': 'Parcours Caisse'
        }
        for old_val, new_val in communes_map.items():
            df.loc[df['Rayon'].str.lower().str.contains(f'^{re.escape(old_val.lower())}$', regex=True, na=False), 'Rayon'] = new_val

    elif report_type in ["Brothers", "Accessories USA", "Accessories Canada"]:
        apply_rule(['GAP', 'GAP Outlet'], ['Acier', 'Deco'], 'Capsules')
        apply_rule(['GAP', 'GAP Outlet'], ['Access'], 'Bijoux')
        apply_rule(['GAP', 'GAP Outlet'], ['Catalog'], 'Enfant')
        communes_map_brothers = {
            'Sacherie': 'Bijoux', 'Catalog': 'Lunettes de vue', 'Papeteri': 'Cartes de voeux',
            'Chaussur': 'Access', 'Mariage': 'Bijoux', 'Montres': 'Bijoux',
            'ParCaiss': 'Parcours Caisse', 'PARCAISS': 'Parcours Caisse'
        }
        for old_val, new_val in communes_map_brothers.items():
            df.loc[df['Rayon'].str.lower().str.contains(f'^{re.escape(old_val.lower())}$', regex=True, na=False), 'Rayon'] = new_val

    return df


# ============================================================
#              LECTURE BLINDÉE DU FICHIER PREV
# ============================================================
MONTH_ORDER = ['janvier', 'février', 'mars', 'avril', 'mai', 'juin',
               'juillet', 'août', 'septembre', 'octobre', 'novembre', 'décembre']


def _find_header_row(df_raw):
    """Cherche la ligne contenant 'Enseigne Client' ET 'Rayon'."""
    for i in range(min(10, len(df_raw))):
        row = df_raw.iloc[i].fillna('').astype(str).str.lower().str.strip().values
        has_enseigne = any('enseigne' in str(v) and 'client' in str(v) for v in row)
        has_rayon = any(str(v).strip() == 'rayon' for v in row)
        if has_enseigne and has_rayon:
            return i
    # Fallback : juste 'rayon' et une autre colonne mois
    for i in range(min(10, len(df_raw))):
        row = df_raw.iloc[i].fillna('').astype(str).str.lower().str.strip().values
        has_rayon = any(str(v).strip() == 'rayon' for v in row)
        has_month = any(remove_accents(str(v)).strip().lower() in [remove_accents(m) for m in MONTH_ORDER] for v in row)
        if has_rayon and has_month:
            return i
    return -1


def _apply_header_row(df_raw, header_idx):
    """Applique la ligne header_idx comme header du DataFrame."""
    raw_headers = df_raw.iloc[header_idx].fillna('').values
    new_cols = [clean_excel_text(str(v)) if str(v).strip() != '' else f'COL_{j}'
                for j, v in enumerate(raw_headers)]
    df = df_raw.iloc[header_idx + 1:].copy()
    df.columns = new_cols
    df = df.reset_index(drop=True)
    # Retirer les lignes entièrement vides
    df = df.dropna(how='all').reset_index(drop=True)
    return df


def read_prev_file(file_buffer):
    """Lecture robuste du fichier Objectifs PREV (CSV ou XLSX).
    Teste séparateurs (; , \\t) et encodages (utf-8, latin1, cp1252) pour les CSV.
    Cherche la ligne contenant 'Enseigne Client' + 'Rayon' comme header."""
    if hasattr(file_buffer, 'seek'):
        file_buffer.seek(0)

    name = (getattr(file_buffer, 'name', '') or '').lower()

    # --- XLSX ---
    if name.endswith('.xlsx') or name.endswith('.xls'):
        try:
            df_raw = pd.read_excel(file_buffer, header=None, dtype=object)
        except Exception as e:
            return None, f"Échec lecture Excel : {e}"
        header_idx = _find_header_row(df_raw)
        if header_idx == -1:
            return None, "Ligne d'en-tête (Enseigne Client + Rayon) introuvable dans le fichier XLSX"
        return _apply_header_row(df_raw, header_idx), None

    # --- CSV : tester toutes les combinaisons ---
    if hasattr(file_buffer, 'seek'):
        file_buffer.seek(0)
    raw_bytes = file_buffer.read()
    if isinstance(raw_bytes, str):
        raw_bytes = raw_bytes.encode('utf-8')

    best_result = None
    best_score = 0

    for encoding in ['utf-8', 'utf-8-sig', 'latin1', 'cp1252']:
        # 1. Décoder et repérer la ligne d'en-tête AVANT read_csv
        try:
            text = raw_bytes.decode(encoding, errors='replace')
        except Exception:
            continue
        lines = text.splitlines()

        header_line_idx = -1
        for i, line in enumerate(lines[:15]):
            low = line.lower()
            if 'enseigne' in low and 'rayon' in low:
                header_line_idx = i
                break
        if header_line_idx == -1:
            continue

        # 2. Tester les séparateurs, en sautant les lignes avant le header
        for sep in [';', ',', '\t', '|']:
            try:
                df = pd.read_csv(
                    io.BytesIO(raw_bytes),
                    sep=sep, encoding=encoding,
                    skiprows=header_line_idx,
                    dtype=str,
                    engine='python',
                    on_bad_lines='skip'
                )
                if df.shape[1] < 5:
                    continue

                # Vérifier présence de 'Enseigne Client' + 'Rayon'
                cols_norm = [remove_accents(str(c)).lower().strip() for c in df.columns]
                has_enseigne = any('enseigne' in c for c in cols_norm)
                has_rayon = any(c == 'rayon' for c in cols_norm)
                if not (has_enseigne and has_rayon):
                    continue

                # Score par nombre de colonnes mois reconnues
                months_norm = [remove_accents(m).lower() for m in MONTH_ORDER]
                score = sum(1 for c in cols_norm if c in months_norm)

                # Nettoyer les noms de colonnes
                df.columns = [clean_excel_text(str(c)) if str(c).strip() != '' else f'COL_{j}'
                              for j, c in enumerate(df.columns)]
                df = df.dropna(how='all').reset_index(drop=True)

                if score > best_score:
                    best_result = df
                    best_score = score
                    if score >= 12:
                        return best_result, None
            except Exception:
                continue

    if best_result is not None:
        return best_result, None

    return None, "Impossible de parser le fichier CSV (aucune combinaison séparateur/encodage n'a fonctionné)"


# ============================================================
#         DÉTECTION DU MOIS DANS LE FICHIER POWER BI
# ============================================================
MONTH_NUM_TO_NAME = {
    1: 'janvier', 2: 'février', 3: 'mars', 4: 'avril',
    5: 'mai', 6: 'juin', 7: 'juillet', 8: 'août',
    9: 'septembre', 10: 'octobre', 11: 'novembre', 12: 'décembre',
}


def detect_month_from_pbi(file_buffer):
    """Lit la première ligne du fichier Power BI pour détecter le mois.
    Cherche des patterns comme 'au 31/03', '31/03/2026', 'mars', etc."""
    if hasattr(file_buffer, 'seek'):
        file_buffer.seek(0)
    try:
        df_peek = pd.read_excel(file_buffer, header=None, nrows=4)
    except Exception:
        return None

    all_text = ' '.join(df_peek.fillna('').astype(str).values.flatten())

    # Pattern 1 : "au JJ/MM" ou "JJ/MM/AAAA"
    m = re.search(r'\b\d{1,2}[/\-](\d{1,2})(?:[/\-]\d{2,4})?\b', all_text)
    if m:
        try:
            month_num = int(m.group(1))
            if 1 <= month_num <= 12:
                return MONTH_NUM_TO_NAME[month_num]
        except ValueError:
            pass

    # Pattern 2 : nom de mois explicite
    all_text_norm = remove_accents(all_text).lower()
    for mnum, mname in MONTH_NUM_TO_NAME.items():
        mname_norm = remove_accents(mname).lower()
        if re.search(rf'\b{mname_norm}\b', all_text_norm):
            return mname

    return None


# ============================================================
#                    CHARGEMENT POWER BI (OBJECTIF)
# ============================================================
def load_powerbi_objectif_data(file_buffer):
    """Charge un export PBI ayant une colonne 'Objectif'."""
    if hasattr(file_buffer, 'seek'):
        file_buffer.seek(0)
    try:
        df_pbi = pd.read_excel(file_buffer, header=1)
    except Exception:
        try:
            if hasattr(file_buffer, 'seek'):
                file_buffer.seek(0)
            df_pbi = pd.read_excel(file_buffer)
        except Exception:
            return pd.DataFrame()

    target_col = 'Objectif'
    if target_col not in df_pbi.columns: return pd.DataFrame()
    if '-Rayon' not in df_pbi.columns or '-Centrale' not in df_pbi.columns: return pd.DataFrame()

    mask_total = df_pbi['-Rayon'].astype(str).str.strip().str.lower() == 'total'
    df_pbi.loc[mask_total, target_col] = df_pbi[target_col].shift(1)
    df_pbi[target_col] = pd.to_numeric(df_pbi[target_col], errors='coerce')
    df_pbi = df_pbi.dropna(subset=[target_col, '-Centrale', '-Rayon']).copy()
    df_pbi = df_pbi[~df_pbi['-Centrale'].astype(str).str.strip().str.lower().isin(['total', 'nan', 'total général'])]
    df_pbi = df_pbi[~df_pbi['-Rayon'].astype(str).str.strip().str.lower().isin(['nan', 'none', ''])]

    df_pbi['-Centrale'] = df_pbi['-Centrale'].apply(clean_excel_text)
    df_pbi['-Rayon'] = df_pbi['-Rayon'].astype(str).apply(clean_excel_text)

    if 'Magasin comparable' in df_pbi.columns:
        df_pbi = df_pbi[~df_pbi['Magasin comparable'].astype(str).str.contains('Dont mag comp', case=False, na=False)]
    df_pbi.loc[df_pbi['-Rayon'].str.lower() == 'total', '-Rayon'] = 'TOTAL'
    df_pbi = df_pbi[~df_pbi['-Centrale'].astype(str).str.contains('Giant Tiger', case=False, na=False)]

    df_pbi['Join_Key'] = df_pbi.apply(lambda x: generate_join_key(x['-Centrale'], x['-Rayon']), axis=1).astype(str)
    df_pbi[target_col] = df_pbi[target_col].apply(safe_float_conversion)

    df_pbi = df_pbi.groupby('Join_Key', as_index=False).agg(
        {'-Centrale': 'first', '-Rayon': 'first', target_col: 'sum'}
    )
    return df_pbi[['Join_Key', '-Centrale', '-Rayon', target_col]]


# ============================================================
#                PIPELINE D'ANALYSE (UN TEST)
# ============================================================
def process_objectif_test(mode, df_prev_source, file_bi, report_type):
    """mode: 'Mensuel' ou 'Annuel'."""
    if file_bi is None:
        return None, None, "Fichier Power BI non fourni"
    if df_prev_source is None:
        return None, None, "Fichier PREV non chargé"

    # Détecter mois depuis BI
    detected_month = detect_month_from_pbi(file_bi)
    if not detected_month:
        return None, None, "Impossible de détecter le mois dans la 1ère ligne du fichier Power BI"

    # Charger PBI
    df_pbi = load_powerbi_objectif_data(file_bi)
    if df_pbi.empty:
        return None, None, "Colonne 'Objectif' introuvable dans le fichier Power BI"

    # Copie PREV pour ne pas la muter
    df_prev = df_prev_source.copy()

    # Mapper les colonnes du PREV aux noms des mois (insensible accents/casse)
    def col_to_month(col):
        c = remove_accents(str(col)).lower().strip()
        for m in MONTH_ORDER:
            if c == remove_accents(m).lower():
                return m
        return None

    month_cols_map = {}
    for c in df_prev.columns:
        m = col_to_month(c)
        if m and m not in month_cols_map:
            month_cols_map[m] = c

    if detected_month not in month_cols_map:
        return None, None, f"Colonne '{detected_month}' introuvable dans le fichier PREV"

    # Calculer col_val selon le mode
    month_idx = MONTH_ORDER.index(detected_month)

    if mode == 'Mensuel':
        cols_to_sum = [month_cols_map[detected_month]]
        diagnostic = f"Mois détecté : **{detected_month.upper()}** · Colonne ciblée dans PREV : {month_cols_map[detected_month]}"
    else:  # Annuel : somme janvier → mois détecté
        cols_to_sum = []
        for i in range(month_idx + 1):
            m = MONTH_ORDER[i]
            if m in month_cols_map:
                cols_to_sum.append(month_cols_map[m])
        diagnostic = f"Mois détecté : **{detected_month.upper()}** · Cumul de {len(cols_to_sum)} mois : {', '.join(cols_to_sum)}"

    # Convertir les colonnes en float
    for c in cols_to_sum:
        df_prev[c] = df_prev[c].apply(safe_float_conversion)

    col_val = 'Valeur_Objectif_PREV'
    df_prev[col_val] = df_prev[cols_to_sum].sum(axis=1)

    # Identifier col_enseigne ("Enseigne Client")
    col_enseigne = None
    for c in df_prev.columns:
        if 'enseigne' in str(c).lower() and 'client' in str(c).lower():
            col_enseigne = c
            break
    if not col_enseigne:
        for c in df_prev.columns:
            if 'enseigne' in str(c).lower() or 'centrale' in str(c).lower():
                col_enseigne = c
                break
    if not col_enseigne:
        return None, None, "Colonne 'Enseigne Client' non détectée dans le fichier PREV"

    # Identifier col_pays
    col_pays = None
    for c in df_prev.columns:
        if str(c).strip().lower() == 'pays':
            col_pays = c
            break

    # Normaliser le nom de la colonne Rayon (apply_business_rules attend 'Rayon')
    for c in df_prev.columns:
        if str(c).strip().lower() == 'rayon' and c != 'Rayon':
            df_prev = df_prev.rename(columns={c: 'Rayon'})
            break

    if 'Rayon' not in df_prev.columns:
        return None, None, "Colonne 'Rayon' non détectée dans le fichier PREV"

    # Exclusion Giant Tiger
    df_prev = df_prev[~df_prev[col_enseigne].astype(str).str.contains('Giant Tiger', case=False, na=False)]

    # Règles métier
    df_prev = apply_business_rules(df_prev, col_enseigne, report_type)

    if report_type == "Diffusion":
        df_prev.loc[df_prev[col_enseigne].astype(str).str.strip().str.upper() == 'GL', col_enseigne] = 'Galeries Lafayette'

    # Filtres géographiques
    if col_pays:
        if report_type == "Diffusion":
            pattern_pays = 'France|Guadeloupe|Guyane|Maurice|Réunion|Reunion|Martinique|Calédonie|Caledonie|Malte|Saint[- ]Martin|Royaume[- ]Uni|Barthelemy|Barthélemy|Luxembourg'
            cond_globale = df_prev[col_pays].astype(str).str.contains(pattern_pays, case=False, na=False, regex=True)
            cond_belgique_ok = df_prev[col_enseigne].astype(str).str.contains('Besson|Paprika|Monoprix|Morgan|Pimkie|Promod', case=False, na=False) & \
                               df_prev[col_pays].astype(str).str.contains('Belgique', case=False, na=False)
            cond_celio_interdit = df_prev[col_enseigne].astype(str).str.contains('Celio', case=False, na=False) & \
                                  ~df_prev[col_pays].astype(str).str.contains('France', case=False, na=False)
            cond_hema_interdit = df_prev[col_enseigne].astype(str).str.contains('Hema', case=False, na=False) & \
                                 ~df_prev[col_pays].astype(str).str.contains('France', case=False, na=False)
            df_prev = df_prev[(cond_globale | cond_belgique_ok) & ~cond_celio_interdit & ~cond_hema_interdit]

            mask_maurice = df_prev[col_pays].astype(str).str.contains('Maurice', case=False, na=False)
            if mask_maurice.sum() > 0:
                df_prev.loc[mask_maurice, col_val] *= 0.019

        elif report_type in ["Brothers", "Accessories USA"]:
            mask_usa = df_prev[col_pays].astype(str).str.contains('USA|Etats-Unis|United States|US', case=False, na=False, regex=True)
            df_prev = df_prev[mask_usa]

        elif report_type == "Accessories Canada":
            mask_canada = df_prev[col_pays].astype(str).str.contains('Canada|CAN', case=False, na=False, regex=True)
            mask_100_canada = df_prev[col_enseigne].astype(str).str.contains('Jean Coutu|Brunet|Red Apple', case=False, na=False)
            df_prev = df_prev[mask_canada | mask_100_canada]

    # Agrégation : par (Enseigne, Rayon) + TOTAL par centrale
    if not df_prev.empty:
        df_prev['Join_Key'] = df_prev.apply(
            lambda x: generate_join_key(x[col_enseigne], x['Rayon']), axis=1
        ).astype(str)
        df_prev_rayons = df_prev.groupby('Join_Key', as_index=False)[col_val].sum()
        df_prev['Centrale_Key'] = df_prev['Join_Key'].apply(lambda x: x.split('_')[0])
        df_prev_total = df_prev.groupby('Centrale_Key', as_index=False)[col_val].sum()
        df_prev_total['Join_Key'] = df_prev_total['Centrale_Key'] + "_TOTAL"
        df_prev_out = pd.concat(
            [df_prev_rayons, df_prev_total[['Join_Key', col_val]]],
            ignore_index=True
        )
        df_prev_out = df_prev_out.groupby('Join_Key', as_index=False)[col_val].sum()
    else:
        df_prev_out = pd.DataFrame(columns=['Join_Key', col_val])

    # Merge PBI ← PREV
    merged = df_pbi.merge(df_prev_out, on='Join_Key', how='left')
    merged['Objectif'] = merged['Objectif'].fillna(0.0)
    merged[col_val] = merged[col_val].fillna(0.0)
    merged['Ecart'] = (merged[col_val] - merged['Objectif']).round(2)
    merged['Statut'] = merged['Ecart'].apply(lambda x: '❌ Anomalie' if abs(x) > 1 else '✅ OK')

    cols_export = ['-Centrale', '-Rayon', col_val, 'Objectif', 'Ecart', 'Statut']
    recap = merged[cols_export].rename(columns={
        '-Centrale': 'Centrale', '-Rayon': 'Rayon',
        col_val: 'Valeur PREV', 'Objectif': 'Objectif Power BI',
    })
    recap = recap.sort_values(by='Statut', ascending=True).reset_index(drop=True)

    return recap, diagnostic, None


def run_full_analysis(file_prev, file_ann, file_mens, report_type, progress_callback=None):
    """Lance les 2 tests (Objectif Annuel, Objectif Mensuel)."""
    all_recaps, diagnostics, skips = {}, {}, {}

    # Lecture PREV
    df_prev_source, prev_err = read_prev_file(file_prev)
    if df_prev_source is None:
        # Tous les tests sont skippés car pas de PREV
        for nom in ["Objectif Annuel", "Objectif Mensuel"]:
            skips[nom] = f"Fichier PREV illisible : {prev_err}"
        return all_recaps, diagnostics, skips

    TESTS = [
        ("Objectif Annuel",  {"mode": "Annuel",  "file": file_ann}),
        ("Objectif Mensuel", {"mode": "Mensuel", "file": file_mens}),
    ]

    for i, (nom_test, conf) in enumerate(TESTS):
        if progress_callback:
            progress_callback(i, len(TESTS), nom_test)
        recap, diag, skip_reason = process_objectif_test(
            conf['mode'], df_prev_source, conf['file'], report_type
        )
        if recap is not None:
            all_recaps[nom_test] = recap
            diagnostics[nom_test] = diag
        elif skip_reason:
            skips[nom_test] = skip_reason

    return all_recaps, diagnostics, skips


# ============================================================
#                    HELPERS POUR L'UI
# ============================================================
def format_fr_number(n):
    try:
        return f"{n:,.2f}".replace(",", " ").replace(".", ",")
    except Exception:
        return str(n)


def tab_label(nom, nb_anomalies):
    return f"✅ {nom}" if nb_anomalies == 0 else f"❌ {nom} ({nb_anomalies})"


def df_to_excel_bytes(df, nom_test):
    safe_name = re.sub(r'[^A-Za-z0-9 _-]', '', nom_test)[:31] or "Recap"
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=safe_name, index=False)
        ws = writer.sheets[safe_name]
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if 'Statut' in df.columns:
            statut_col_idx = list(df.columns).index('Statut') + 1
            for row_idx in range(2, len(df) + 2):
                statut_val = ws.cell(row=row_idx, column=statut_col_idx).value
                fill = red_fill if statut_val and 'Anomalie' in str(statut_val) else green_fill
                for col_idx in range(1, len(df.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = max(len(str(col_name)),
                          df[col_name].astype(str).str.len().max() if len(df) > 0 else 10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    return output.getvalue()


def style_dataframe(df):
    def highlight_row(row):
        if row['Statut'] == '❌ Anomalie':
            return ['background-color: #fef2f2; color: #991b1b'] * len(row)
        return ['background-color: #f0fdf4; color: #166534'] * len(row)
    return df.style.apply(highlight_row, axis=1).format({
        'Valeur PREV': lambda x: format_fr_number(x),
        'Objectif Power BI': lambda x: format_fr_number(x),
        'Ecart': lambda x: format_fr_number(x),
    })


# ============================================================
#                    SESSION STATE INIT
# ============================================================
for key, default in [
    ('results', None), ('diagnostics', {}), ('skips', {}),
    ('last_run_type', None), ('last_run_time', None),
]:
    if key not in st.session_state:
        st.session_state[key] = default


# ============================================================
#                       SIDEBAR
# ============================================================
with st.sidebar:
    st.markdown('<div class="sidebar-title">⚙️ Configuration</div>', unsafe_allow_html=True)
    report_type = st.selectbox(
        "Type de reporting",
        ["Diffusion", "Brothers", "Accessories USA", "Accessories Canada"],
        help="Détermine les règles métier et filtres géographiques appliqués."
    )

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<div class="sidebar-title">📎 Fichiers</div>', unsafe_allow_html=True)

    file_prev = st.file_uploader("Fichier Objectifs PREV (obligatoire)", type=['xlsx', 'xls', 'csv'])
    if file_prev:
        st.markdown(f'<div class="upload-ok">✓ {file_prev.name}</div>', unsafe_allow_html=True)

    file_ann = st.file_uploader("Export BI — Annuel", type=['xlsx'])
    if file_ann:
        st.markdown(f'<div class="upload-ok">✓ {file_ann.name}</div>', unsafe_allow_html=True)

    file_mens = st.file_uploader("Export BI — Mensuel", type=['xlsx'])
    if file_mens:
        st.markdown(f'<div class="upload-ok">✓ {file_mens.name}</div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    bi_loaded = sum(1 for f in [file_ann, file_mens] if f is not None)
    launch_btn = st.button(
        "🚀 Lancer l'Analyse",
        type="primary",
        disabled=(file_prev is None or bi_loaded == 0),
        use_container_width=True
    )

    if file_prev is None:
        st.caption("⚠️ Importe le fichier PREV.")
    elif bi_loaded == 0:
        st.caption("⚠️ Importe au moins un export Power BI.")
    else:
        st.caption(f"✓ Prêt · {bi_loaded} fichier(s) BI chargé(s)")

    st.markdown("---")
    with st.expander("ℹ️ Aide & Documentation"):
        st.markdown("""
**Fichiers attendus**
- **PREV** (`.xlsx` ou `.csv`) : colonnes `Enseigne Client`, `Rayon`, `Pays`, `janvier`…`décembre`.
  La ligne 1 peut contenir des commentaires, ils sont ignorés automatiquement.
- **BI Annuel / Mensuel** : exports Power BI avec colonnes `-Centrale`, `-Rayon`, **`Objectif`**.
  La 1ère ligne doit mentionner le mois en cours (ex: `au 31/03` → mars).

**Calcul de la valeur PREV**
- *Objectif Mensuel* : la colonne du mois détecté uniquement.
- *Objectif Annuel* : somme des colonnes de janvier au mois détecté inclus.

**Tolérance d'écart**
Un écart supérieur à **1 €** est classé anomalie.
""")

    if st.session_state.last_run_time:
        st.caption(f"Dernière analyse : {st.session_state.last_run_time}")


# ============================================================
#                      ZONE PRINCIPALE
# ============================================================
st.markdown('<div class="main-title">🎯 Réconciliation Objectifs PREV vs Power BI</div>', unsafe_allow_html=True)
st.markdown(
    f'<div class="main-subtitle">Type de reporting : <b>{report_type}</b> '
    f'· Contrôle automatisé des écarts entre objectifs PREV et Power BI</div>',
    unsafe_allow_html=True
)

# Lancement de l'analyse
if launch_btn and file_prev:
    progress_placeholder = st.empty()
    progress_bar = progress_placeholder.progress(0.0, text="Initialisation...")

    def update_progress(current, total, test_name):
        pct = (current + 1) / total
        progress_bar.progress(pct, text=f"[{current+1}/{total}] Analyse : {test_name}")

    try:
        all_recaps, diagnostics, skips = run_full_analysis(
            file_prev, file_ann, file_mens, report_type,
            progress_callback=update_progress
        )
        st.session_state.results = all_recaps
        st.session_state.diagnostics = diagnostics
        st.session_state.skips = skips
        st.session_state.last_run_type = report_type
        st.session_state.last_run_time = datetime.now().strftime("%d/%m/%Y à %H:%M")
        progress_placeholder.empty()
    except Exception as e:
        progress_placeholder.empty()
        st.error(f"❌ Erreur durant l'analyse : {e}")
        st.session_state.results = None

# Affichage résultats
if st.session_state.results:
    all_recaps = st.session_state.results
    diagnostics = st.session_state.diagnostics
    skips = st.session_state.skips

    if not all_recaps:
        st.warning("Aucun résultat n'a pu être généré. Vérifie les fichiers et les onglets.")
        if skips:
            with st.expander("🔍 Détails des tests ignorés", expanded=True):
                for nom, raison in skips.items():
                    st.markdown(f"- **{nom}** : {raison}")
    else:
        # --- KPIs ---
        total_tests = len(all_recaps)
        total_lignes = sum(len(df) for df in all_recaps.values())
        total_anomalies = sum(len(df[df['Statut'] == '❌ Anomalie']) for df in all_recaps.values())
        total_ok = total_lignes - total_anomalies

        k1, k2, k3, k4 = st.columns(4)
        with k1:
            st.markdown(f"""<div class="kpi-card">
                <div class="kpi-label">Tests effectués</div>
                <div class="kpi-value">{total_tests}</div></div>""", unsafe_allow_html=True)
        with k2:
            st.markdown(f"""<div class="kpi-card">
                <div class="kpi-label">Lignes vérifiées</div>
                <div class="kpi-value">{total_lignes:,}</div></div>""".replace(",", " "), unsafe_allow_html=True)
        with k3:
            st.markdown(f"""<div class="kpi-card kpi-card-ok">
                <div class="kpi-label">Lignes OK</div>
                <div class="kpi-value kpi-value-ok">{total_ok:,}</div></div>""".replace(",", " "), unsafe_allow_html=True)
        with k4:
            alert_class = "kpi-card-ok" if total_anomalies == 0 else "kpi-card-alert"
            value_class = "kpi-value-ok" if total_anomalies == 0 else "kpi-value-alert"
            st.markdown(f"""<div class="kpi-card {alert_class}">
                <div class="kpi-label">Anomalies</div>
                <div class="kpi-value {value_class}">{total_anomalies:,}</div></div>""".replace(",", " "), unsafe_allow_html=True)

        if skips:
            with st.expander(f"⚠️ {len(skips)} test(s) ignoré(s)"):
                for nom, raison in skips.items():
                    st.markdown(f"- **{nom}** : {raison}")

        st.markdown('<div class="section-header">📋 Résultats détaillés par test</div>', unsafe_allow_html=True)

        # Tri par gravité
        sorted_items = sorted(
            all_recaps.items(),
            key=lambda x: -len(x[1][x[1]['Statut'] == '❌ Anomalie'])
        )
        labels = [tab_label(nom, len(df[df['Statut'] == '❌ Anomalie'])) for nom, df in sorted_items]
        tabs = st.tabs(labels)

        for tab, (nom, df) in zip(tabs, sorted_items):
            with tab:
                if nom in diagnostics and diagnostics[nom]:
                    st.markdown(
                        f'<div class="diag-box">🔍 <b>Diagnostic :</b> {diagnostics[nom]}</div>',
                        unsafe_allow_html=True
                    )

                nb_anomalies = len(df[df['Statut'] == '❌ Anomalie'])
                nb_total = len(df)

                a1, a2, a3 = st.columns([5, 2, 2])
                with a1:
                    search = st.text_input(
                        "Rechercher",
                        placeholder="🔎 Rechercher par centrale ou rayon...",
                        key=f"search_{nom}",
                        label_visibility="collapsed"
                    )
                with a2:
                    only_anom = st.toggle(
                        "Anomalies seulement",
                        value=(nb_anomalies > 0),
                        key=f"toggle_{nom}"
                    )
                with a3:
                    try:
                        excel_bytes = df_to_excel_bytes(df, nom)
                        st.download_button(
                            "📥 Télécharger",
                            data=excel_bytes,
                            file_name=f"recap_{nom.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_{nom}",
                            use_container_width=True
                        )
                    except Exception:
                        st.caption("Export indisponible")

                df_view = df.copy()
                if only_anom:
                    df_view = df_view[df_view['Statut'] == '❌ Anomalie']
                if search:
                    s = search.lower()
                    mask = (
                        df_view['Centrale'].astype(str).str.lower().str.contains(s, na=False) |
                        df_view['Rayon'].astype(str).str.lower().str.contains(s, na=False)
                    )
                    df_view = df_view[mask]

                if nb_anomalies == 0:
                    st.success(f"✅ Tout est OK sur cet onglet — **{nb_total}** ligne(s) vérifiée(s)")
                else:
                    if only_anom:
                        st.error(f"❌ **{nb_anomalies}** anomalie(s) · {len(df_view)} affichée(s)")
                    else:
                        st.warning(f"⚠️ **{nb_anomalies}** anomalie(s) sur {nb_total} ligne(s) · {len(df_view)} affichée(s)")

                if len(df_view) == 0:
                    st.info("Aucune ligne ne correspond aux filtres.")
                else:
                    try:
                        st.dataframe(
                            style_dataframe(df_view),
                            use_container_width=True,
                            hide_index=True,
                            height=min(600, 60 + 35 * min(len(df_view), 15))
                        )
                    except Exception:
                        st.dataframe(df_view, use_container_width=True, hide_index=True)

elif not file_prev:
    st.info(
        "👋 **Bienvenue !** Importe ton fichier Objectifs PREV et au moins un export Power BI "
        "dans la barre latérale à gauche, puis clique sur **🚀 Lancer l'Analyse**."
    )
    with st.expander("📖 Comment ça marche ?", expanded=False):
        st.markdown("""
**1. Prépare tes fichiers**
- Le fichier **PREV** (xlsx ou csv) avec les objectifs par mois (colonnes `janvier`…`décembre`)
- Un ou deux exports Power BI : Annuel et/ou Mensuel (avec une colonne `Objectif`)

**2. Choisis le type de reporting**
Les règles métier et filtres pays s'adaptent :
- **Diffusion** : France + DOM-TOM + Europe, avec règles Maurice/Belgique/Celio/Hema
- **Brothers / USA** : ciblage Etats-Unis
- **Canada** : Canada + enseignes 100% canadiennes (Jean Coutu, Brunet, Red Apple)

**3. Lance l'analyse**
L'outil détecte automatiquement le mois en cours dans la 1ère ligne du fichier BI (ex: `au 31/03` → mars).

**4. Examine les résultats**
- **Objectif Mensuel** : écart entre la colonne du mois PREV et la colonne Objectif du BI Mensuel
- **Objectif Annuel** : écart entre le cumul janvier→mois détecté du PREV et la colonne Objectif du BI Annuel
""")

else:
    st.success("✅ Fichiers prêts. Clique sur **🚀 Lancer l'Analyse** dans la barre latérale.")